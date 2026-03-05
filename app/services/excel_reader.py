from __future__ import annotations

from datetime import date
from pathlib import Path
from typing import Any, Dict, List

import openpyxl


MaintenanceRow = Dict[str, str]


def _normalize(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _compact_header(text: str) -> str:
    return "".join(text.split())


def _find_processing_date_key(headers: List[str]) -> str | None:
    for header in headers:
        if _compact_header(header) == "\ucc98\ub9ac\uc77c\uc790":
            return header
    return None


def _find_status_column_index(headers: List[str]) -> int | None:
    for idx, header in enumerate(headers, start=1):
        if _compact_header(header) == "\uc9c4\ud589\uc0c1\ud0dc":
            return idx
    return None


def _is_completed_status(value: str) -> bool:
    return _compact_header(_normalize(value)) == "\uc644\ub8cc"


def _parse_date_value(text: str) -> date | None:
    s = _normalize(text)
    if not s:
        return None

    # Accept common forms: YYYY-MM-DD, YYYY/MM/DD, YYYY.MM.DD, YYYY M D
    digits: List[int] = []
    current = ""
    for ch in s:
        if ch.isdigit():
            current += ch
        elif current:
            digits.append(int(current))
            current = ""
    if current:
        digits.append(int(current))

    if len(digits) < 3:
        return None

    y, m, d = digits[0], digits[1], digits[2]
    if y < 100:
        y += 2000

    try:
        return date(y, m, d)
    except ValueError:
        return None


def read_maintenance_rows(xlsx_path: Path, required_headers: List[str]) -> List[MaintenanceRow]:
    """Read rows from Excel using template placeholders and matched headers."""
    required: List[str] = []
    for name in required_headers:
        n = _normalize(name)
        if n and n not in required:
            required.append(n)
    if not required:
        raise ValueError("No template placeholders were provided.")

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    header_row = None
    headers: List[str] = []
    best_overlap = 0
    for r in range(1, min(ws.max_row, 20) + 1):
        values = [_normalize(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
        overlap = sum(1 for h in required if h in values)
        if overlap > best_overlap:
            best_overlap = overlap
            header_row = r
            headers = values
        if overlap == len(required):
            break

    if header_row is None or best_overlap == 0:
        raise ValueError("Could not find a header row that matches template placeholders.")

    col_index: Dict[str, int] = {}
    for idx, name in enumerate(headers, start=1):
        if name in required:
            col_index[name] = idx

    rows: List[MaintenanceRow] = []
    status_col_index = _find_status_column_index(headers)
    if status_col_index is None:
        raise ValueError("Missing required Excel column for filtering: 진행상태")

    for r in range(header_row + 1, ws.max_row + 1):
        row_data: MaintenanceRow = {header: "" for header in required}
        for header in required:
            idx = col_index.get(header)
            if idx is not None:
                row_data[header] = _normalize(ws.cell(r, idx).value)

        if not any(row_data.values()):
            continue

        status_value = _normalize(ws.cell(r, status_col_index).value)
        if not _is_completed_status(status_value):
            continue

        rows.append(row_data)

    processing_date_key = _find_processing_date_key(required)
    if processing_date_key:
        rows.sort(
            key=lambda row: (
                _parse_date_value(row.get(processing_date_key, "")) is None,
                _parse_date_value(row.get(processing_date_key, "")) or date.max,
            )
        )

    return rows
